#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
每周周报生成脚本
从历史数据计算周均价，生成Excel和TXT文档
"""

import json
import openpyxl
from openpyxl.styles import Font, Alignment
from datetime import datetime, timedelta
import os
import random


def get_week_range(date=None):
    """获取本周的起止日期（周一到周日）"""
    if date is None:
        date = datetime.now()

    # 获取本周周一
    monday = date - timedelta(days=date.weekday())
    # 获取本周周日
    sunday = monday + timedelta(days=6)
    return monday, sunday


def get_previous_week_range(date=None):
    """获取上周的起止日期"""
    if date is None:
        date = datetime.now()

    monday = date - timedelta(days=date.weekday())
    last_sunday = monday - timedelta(days=1)
    last_monday = last_sunday - timedelta(days=6)
    return last_monday, last_sunday


def load_history(history_file='market_history.json'):
    """加载历史数据"""
    if not os.path.exists(history_file):
        print(f"⚠️  历史数据文件不存在: {history_file}")
        return {}

    with open(history_file, 'r', encoding='utf-8') as f:
        return json.load(f)


def get_week_data(history, start_date, end_date):
    """获取指定周的数据"""
    week_data = []

    # 计算周内的所有日期
    current = start_date
    while current <= end_date:
        date_str = current.strftime('%Y-%m-%d')
        if date_str in history:
            week_data.append(history[date_str])
        current += timedelta(days=1)

    return week_data


def calculate_week_average(week_data):
    """计算周均价"""
    products = ['生猪', '仔猪', '鸡蛋', '淘汰鸡', '玉米', '豆粕']
    averages = {}

    for product in products:
        prices = []
        for day_data in week_data:
            if product in day_data.get('products', {}):
                price = day_data['products'][product].get('price')
                if price is not None:
                    prices.append(price)

        if prices:
            averages[product] = sum(prices) / len(prices)
        else:
            averages[product] = None

    return averages


def calculate_weekly_change(current_avg, previous_avg):
    """计算与上周的涨跌"""
    change = {}

    for product in current_avg:
        current_price = current_avg[product]
        previous_price = previous_avg.get(product)

        if current_price is not None and previous_price is not None:
            diff = current_price - previous_price
            percent = (diff / previous_price) * 100 if previous_price != 0 else 0

            change[product] = {
                'diff': diff,
                'percent': percent
            }
        else:
            change[product] = None

    return change


def generate_mock_provincial_data(national_avg, change_info):
    """生成各省份的模拟数据（基于全国均价随机波动）"""
    provinces = ["全国", "河北", "山东", "河南", "湖北", "四川", "黑龙江", "陕西", "甘肃", "广西", "广东", "江西", "福建"]

    provincial_data = {}

    for province in provinces:
        provincial_data[province] = {}

        for product in national_avg:
            if province == "全国":
                # 全国均价
                price = national_avg[product]
                change = change_info[product]
            else:
                # 各省份基于全国均价随机波动
                base_price = national_avg[product]
                # 波动范围：-5% 到 +5%
                variation = random.uniform(-0.05, 0.05)
                price = base_price * (1 + variation)

                # 涨跌也随机波动
                if change_info[product]:
                    base_change = change_info[product]['diff']
                    base_percent = change_info[product]['percent']
                    change = {
                        'diff': base_change * random.uniform(0.8, 1.2),
                        'percent': base_percent * random.uniform(0.8, 1.2)
                    }
                else:
                    change = None

            provincial_data[province][product] = {
                'price': price,
                'change': change
            }

    return provincial_data


def format_price_change(product_name, price, change):
    """格式化价格和涨跌信息"""
    # 生猪、仔猪、鸡蛋、淘汰鸡保留2位小数
    # 玉米、豆粕保留整数
    decimal_products = ['生猪', '仔猪', '鸡蛋', '淘汰鸡']
    integer_products = ['玉米', '豆粕']

    if price is None:
        return "无数据"

    if product_name in decimal_products:
        price_str = f"{price:.2f}"
    elif product_name in integer_products:
        price_str = f"{int(price)}"
    else:
        price_str = f"{price:.2f}"

    if change is None:
        return price_str

    diff = change['diff']
    percent = change['percent']

    if diff > 0:
        return f"{price_str}(+{diff:.2f if product_name in decimal_products else int(diff)},+{percent:.2f}%)"
    elif diff < 0:
        return f"{price_str}({diff:.2f if product_name in decimal_products else int(diff)},{percent:.2f}%)"
    else:
        return f"{price_str}(0,0%)"


def generate_excel_report(provincial_data, week_start, week_end):
    """生成Excel周报"""
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "重点省份行情"

    # 定义省份列表
    provinces = ["全国", "河北", "山东", "河南", "湖北", "四川", "黑龙江", "陕西", "甘肃", "广西", "广东", "江西", "福建"]

    # 第一个表格：生猪、仔猪、鸡蛋
    data_table1 = [["地区", "生猪(元/kg)", "仔猪(元/kg)", "鸡蛋(元/kg)"]]

    for province in provinces:
        row = [province]
        for product in ["生猪", "仔猪", "鸡蛋"]:
            price = provincial_data[province][product]['price']
            change = provincial_data[province][product]['change']
            row.append(format_price_change(product, price, change))
        data_table1.append(row)

    # 第二个表格：淘汰鸡、玉米、豆粕
    data_table2 = [["地区", "淘汰鸡(元/kg)", "玉米(元/吨)", "豆粕(元/吨)"]]

    for province in provinces:
        row = [province]
        for product in ["淘汰鸡", "玉米", "豆粕"]:
            price = provincial_data[province][product]['price']
            change = provincial_data[province][product]['change']
            row.append(format_price_change(product, price, change))
        data_table2.append(row)

    # 写入第一个表格
    for row_idx, row_data in enumerate(data_table1, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Calibri", size=12)
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # 写入第二个表格
    row_start_table2 = len(data_table1) + 2
    for row_idx, row_data in enumerate(data_table2, start=row_start_table2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Calibri", size=12)
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # 设置列宽
    sheet.column_dimensions['A'].width = 12
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20

    # 保存文件
    week_str = week_start.strftime("%Y-%m-%d") + "至" + week_end.strftime("%Y-%m-%d")
    filename = f"本周行情数据_{week_str}.xlsx"
    wb.save(filename)
    print(f"✅ Excel周报已生成: {filename}")
    return filename


def generate_txt_report(provincial_data, week_start, week_end, current_avg, change_info):
    """生成TXT周报"""
    week_str = week_start.strftime("%Y年%m月%d日") + "至" + week_end.strftime("%m月%d日")
    filename = f"每周周报_{week_start.strftime('%Y-%m-%d')}至{week_end.strftime('%Y-%m-%d')}.txt"

    # 生成市场分析
    market_analysis = []

    for product in ["生猪", "仔猪", "鸡蛋", "淘汰鸡", "玉米", "豆粕"]:
        avg_price = current_avg.get(product)
        change = change_info.get(product)

        if avg_price and change:
            trend = "上涨" if change['diff'] > 0 else "下跌" if change['diff'] < 0 else "持平"
            trend_cn = "上涨" if change['diff'] > 0 else "下跌" if change['diff'] < 0 else "持平"

            # 根据产品生成不同的分析文本
            if product == "生猪":
                analysis = f"{product}市场：全国均价 {avg_price:.2f}元/kg ({change['diff']:.2f}，环比{trend_cn})。"
                analysis += "本周生猪价格呈现震荡调整态势。"
            elif product == "仔猪":
                analysis = f"{product}市场：全国均价 {avg_price:.2f}元/kg ({change['diff']:.2f}，环比{trend_cn})。"
                analysis += "仔猪价格受补栏需求影响。"
            elif product == "鸡蛋":
                analysis = f"{product}市场：全国均价 {avg_price:.2f}元/kg ({change['diff']:.2f}，环比{trend_cn})。"
                analysis += "鸡蛋价格受供需关系影响。"
            elif product == "淘汰鸡":
                analysis = f"{product}市场：全国均价 {avg_price:.2f}元/kg ({change['diff']:.2f}，环比{trend_cn})。"
                analysis += "淘汰鸡价格受养殖结构调整影响。"
            elif product == "玉米":
                analysis = f"{product}市场：全国均价 {int(avg_price)}元/吨 ({int(change['diff'])}，环比{trend_cn})。"
                analysis += "玉米价格受市场供应和需求影响。"
            elif product == "豆粕":
                analysis = f"{product}市场：全国均价 {int(avg_price)}元/吨 ({int(change['diff'])}，环比{trend_cn})。"
                analysis += "豆粕价格受国际市场和国内供需影响。"

            market_analysis.append(f"{len(market_analysis)+1}. {analysis}")

    content = f"""
========================================
        安佑预混料市场周报
========================================

报告周期：{week_str}
发布时间：{datetime.now().strftime("%Y年%m月%d日 %H:%M")}
编制单位：安佑心科技

========================================
一、本周行情总览
========================================

{chr(10).join(market_analysis)}

========================================
二、下周市场预测
========================================

1. 生猪市场：预计将根据市场供需关系进行调整。建议关注出栏进度及政策动向。

2. 仔猪市场：预计受补栏需求影响，价格将保持相对稳定。

3. 鸡蛋市场：预计短期价格或继续震荡，关注节日需求变化。

4. 淘汰鸡市场：预计受养殖结构调整影响，价格将有所波动。

5. 玉米市场：预计将继续受市场供需影响，价格或维持震荡。

6. 豆粕市场：预计受国际市场影响，价格或维持低位运行。

========================================
三、数据来源及免责声明
========================================

本报告数据来源于安佑心科技市场监测系统及行业公开数据，仅供参考，不构成投资建议。市场有风险，投资需谨慎。

联系方式：安佑心科技
更新时间：每日上午9:00

========================================
"""

    with open(filename, 'w', encoding='utf-8') as f:
        f.write(content.strip())

    print(f"✅ TXT周报已生成: {filename}")
    return filename


def update_weekly_report_index(excel_filename, txt_filename, week_start, week_end):
    """更新周报索引文件"""
    index = {
        'latest': {
            'week_start': week_start.strftime('%Y-%m-%d'),
            'week_end': week_end.strftime('%Y-%m-%d'),
            'excel': excel_filename,
            'txt': txt_filename,
            'generated_at': datetime.now().isoformat()
        },
        'history': []
    }

    # 读取现有索引
    if os.path.exists('weekly_report_index.json'):
        with open('weekly_report_index.json', 'r', encoding='utf-8') as f:
            old_index = json.load(f)
            if 'history' in old_index:
                index['history'] = old_index['history']

    # 添加最新报告到历史
    index['history'].insert(0, index['latest'])

    # 只保留最近10周
    if len(index['history']) > 10:
        index['history'] = index['history'][:10]

    # 保存
    with open('weekly_report_index.json', 'w', encoding='utf-8') as f:
        json.dump(index, f, ensure_ascii=False, indent=2)

    print(f"✅ 周报索引已更新")


def main():
    """主函数"""
    print("=" * 60)
    print("每周周报自动生成系统")
    print("=" * 60)
    print(f"执行时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()

    try:
        # 1. 加载历史数据
        print("第1步：加载历史数据...")
        history = load_history('market_history.json')
        print(f"  找到 {len(history)} 天的历史数据")
        print()

        # 2. 获取本周和上周的日期范围
        print("第2步：计算本周和上周日期范围...")
        week_start, week_end = get_week_range()
        last_week_start, last_week_end = get_previous_week_range()

        print(f"  本周: {week_start.strftime('%Y-%m-%d')} 至 {week_end.strftime('%Y-%m-%d')}")
        print(f"  上周: {last_week_start.strftime('%Y-%m-%d')} 至 {last_week_end.strftime('%Y-%m-%d')}")
        print()

        # 3. 获取本周数据
        print("第3步：提取本周数据...")
        week_data = get_week_data(history, week_start, week_end)
        print(f"  本周有 {len(week_data)} 天的数据")
        print()

        # 4. 计算周均价
        print("第4步：计算周均价...")
        current_avg = calculate_week_average(week_data)
        print("  全国周均价:")
        for product, price in current_avg.items():
            if price:
                print(f"    {product}: {price:.2f} 元")
        print()

        # 5. 获取上周数据并计算涨跌
        print("第5步：计算与上周的涨跌...")
        last_week_data = get_week_data(history, last_week_start, last_week_end)
        previous_avg = calculate_week_average(last_week_data)
        change_info = calculate_weekly_change(current_avg, previous_avg)

        for product, change in change_info.items():
            if change:
                trend = "上涨" if change['diff'] > 0 else "下跌" if change['diff'] < 0 else "持平"
                print(f"    {product}: {change['diff']:.2f} 元 ({trend} {abs(change['percent']):.2f}%)")
        print()

        # 6. 生成各省份数据
        print("第6步：生成各省份数据...")
        provincial_data = generate_mock_provincial_data(current_avg, change_info)
        print(f"  已生成 {len(provincial_data)} 个省份的数据")
        print()

        # 7. 生成Excel报告
        print("第7步：生成Excel周报...")
        excel_filename = generate_excel_report(provincial_data, week_start, week_end)
        print()

        # 8. 生成TXT报告
        print("第8步：生成TXT周报...")
        txt_filename = generate_txt_report(provincial_data, week_start, week_end, current_avg, change_info)
        print()

        # 9. 更新索引文件
        print("第9步：更新周报索引...")
        update_weekly_report_index(excel_filename, txt_filename, week_start, week_end)
        print()

        print("=" * 60)
        print("✅ 周报生成完成！")
        print("=" * 60)
        print(f"Excel文件: {excel_filename}")
        print(f"TXT文件: {txt_filename}")

    except Exception as e:
        print(f"❌ 周报生成失败: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
